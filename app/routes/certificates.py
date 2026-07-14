from flask import Blueprint, render_template, request, jsonify, current_app
from flask_login import login_required, current_user
from app.models import (db, User, CertificateType, AuditLog,
                         CertArchive, EmailDraft, OrgSettings, Campaign)
from app.engine.cert_id import generate_cert_id, next_sequence, verify_format
from app.engine.ocr_analyzer import analyze_template
from datetime import datetime
import uuid, os, json, io, csv

bp = Blueprint('certificates', __name__)
UPLOAD_FOLDER = 'uploads'

@bp.route('/admin')
@login_required
def dashboard():
    cert_types = CertificateType.query.filter_by(is_active=True)\
        .order_by(CertificateType.created_at.desc()).all()
    org = OrgSettings.query.first() or OrgSettings()
    drafts = EmailDraft.query.order_by(EmailDraft.updated_at.desc()).all()
    return render_template('admin/dashboard.html',
                           cert_types=cert_types, org=org, drafts=drafts)

@bp.route('/api/org', methods=['GET'])
@login_required
def get_org():
    org = OrgSettings.query.first() or OrgSettings()
    return jsonify({
        'org_name': org.org_name,
        'sender_name': org.sender_name,
        'sender_email': org.sender_email,
        'reply_to_email': org.reply_to_email,
        'verify_base_url': org.verify_base_url,
    })


@bp.route('/api/org', methods=['POST'])
@login_required
def save_org():
    data = request.json or {}
    org = OrgSettings.query.first()
    if not org:
        org = OrgSettings()
        db.session.add(org)
    for field in ('org_name', 'sender_name', 'sender_email', 'reply_to_email', 'verify_base_url'):
        if field in data:
            setattr(org, field, data[field])
    db.session.commit()
    return jsonify({'message': 'Settings saved'})

@bp.route('/api/cert-types', methods=['GET'])
@login_required
def list_cert_types():
    types = CertificateType.query.filter_by(is_active=True)\
        .order_by(CertificateType.created_at.desc()).all()
    return jsonify([{
        'id': ct.id, 'name': ct.name, 'course_code': ct.course_code,
        'period': ct.period, 'registration_token': ct.registration_token,
        'master_file_type': ct.master_file_type,
        'render_mode': ct.render_mode or 'mlj',
        'user_count': len(ct.users),
        'seq_counter': ct.seq_counter,
    } for ct in types])


@bp.route('/api/cert-types', methods=['POST'])
@login_required
def create_cert_type():
    """Create a certificate gateway. Uploading a master template is now
    optional: with a file the legacy overlay engine is used; without one the
    gateway renders the MedLocum designed certificate (ported from the
    medlocumjobs e-learning pipeline) — no coordinates or OCR needed."""
    file = request.files.get('master_file')
    name = request.form.get('name', '').strip()
    period = request.form.get('period', '').strip()
    course_code = request.form.get('course_code', 'GEN').strip().upper()
    if not name or not period:
        return jsonify({'error': 'Name and period required'}), 400

    token = str(uuid.uuid4())[:8]
    ocr_result = {'regions': None, 'message': None}

    if file:
        ext = file.filename.rsplit('.', 1)[-1].lower()
        if ext not in ('pdf', 'png'):
            return jsonify({'error': 'Only PDF or PNG allowed'}), 400

        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        filename = f"{uuid.uuid4().hex}.{ext}"
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)

        overlay_coords = {
            "name_x": 0, "name_y": 320, "name_w": 500, "name_h": 50,
            "name_font_size": 30, "name_align": "center",
            "cert_id_x": 60, "cert_id_y": 55, "cert_id_font_size": 9,
            "date_x": 60, "date_y": 42, "date_font_size": 9,
            "qr_x": -1, "qr_y": 22, "qr_size": 70
        }
        try:
            custom = json.loads(request.form.get('overlay_coords') or 'null')
            if custom:
                overlay_coords.update(custom)
        except Exception:
            pass

        try:
            ocr_result = analyze_template(filepath, ext)
        except Exception as ocr_err:
            ocr_result = {'regions': [], 'message': f'OCR skipped: {ocr_err}'}

        ct = CertificateType(
            name=name, course_code=course_code, period=period,
            render_mode='overlay',
            master_pdf_path=filepath, master_file_type=ext,
            overlay_coords=overlay_coords,
            ocr_regions=ocr_result.get('regions'),
            registration_token=token,
            email_message=request.form.get('email_message', '').strip() or None,
            email_subject=request.form.get('email_subject', '').strip() or None,
            seq_counter=0,
        )
    else:
        def _float(value, default=0.0):
            try:
                return float(value)
            except (TypeError, ValueError):
                return default

        ct = CertificateType(
            name=name, course_code=course_code, period=period,
            render_mode='mlj',
            master_pdf_path='', master_file_type='',
            overlay_coords={},
            registration_token=token,
            organisation=request.form.get('organisation', '').strip(),
            website=request.form.get('website', '').strip(),
            signatory_name=request.form.get('signatory_name', '').strip(),
            signatory_title=request.form.get('signatory_title', '').strip(),
            duration_hours=_float(request.form.get('duration_hours')),
            email_message=request.form.get('email_message', '').strip() or None,
            email_subject=request.form.get('email_subject', '').strip() or None,
            seq_counter=0,
        )

    db.session.add(ct)
    db.session.commit()

    return jsonify({
        'id': ct.id, 'name': ct.name, 'registration_token': token,
        'render_mode': ct.render_mode,
        'register_url': f"/register/{token}",
        'ocr_message': ocr_result.get('message'),
        'ocr_regions': ocr_result.get('regions'),
    })


@bp.route('/api/cert-types/<int:ct_id>', methods=['GET'])
@login_required
def get_cert_type(ct_id):
    ct = CertificateType.query.get_or_404(ct_id)
    return jsonify({
        'id': ct.id, 'name': ct.name, 'course_code': ct.course_code,
        'period': ct.period, 'master_file_type': ct.master_file_type,
        'render_mode': ct.render_mode or 'mlj',
        'overlay_coords': ct.overlay_coords,
        'ocr_regions': ct.ocr_regions,
        'organisation': ct.organisation,
        'website': ct.website,
        'signatory_name': ct.signatory_name,
        'signatory_title': ct.signatory_title,
        'duration_hours': ct.duration_hours,
        'email_subject': ct.email_subject,
        'email_message': ct.email_message,
        'registration_token': ct.registration_token,
    })


@bp.route('/api/cert-types/<int:ct_id>', methods=['PUT'])
@login_required
def update_cert_type(ct_id):
    ct = CertificateType.query.get_or_404(ct_id)
    data = request.json or {}
    for field in ('name', 'period', 'course_code', 'overlay_coords',
                  'email_message', 'email_subject', 'ocr_regions',
                  'organisation', 'website', 'signatory_name',
                  'signatory_title', 'duration_hours'):
        if field in data:
            setattr(ct, field, data[field])
    db.session.commit()
    return jsonify({'message': 'Updated'})


@bp.route('/api/cert-types/<int:ct_id>', methods=['DELETE'])
@login_required
def delete_cert_type(ct_id):
    ct = CertificateType.query.get_or_404(ct_id)
    ct.is_active = False
    db.session.commit()
    return jsonify({'message': 'Removed from library'})

@bp.route('/api/cert-types/<int:ct_id>/analyze', methods=['POST'])
@login_required
def reanalyze_template(ct_id):
    ct = CertificateType.query.get_or_404(ct_id)
    result = analyze_template(ct.master_pdf_path, ct.master_file_type)
    ct.ocr_regions = result.get('regions')
    db.session.commit()
    return jsonify({'regions': ct.ocr_regions, 'message': result.get('message')})


@bp.route('/api/cert-types/<int:ct_id>/preview', methods=['POST'])
@login_required
def preview_certificate(ct_id):
    from flask import send_file
    ct = CertificateType.query.get_or_404(ct_id)
    data = request.json or {}
    full_name = data.get('full_name', 'Jane Smith')
    cert_id = data.get('cert_id', 'MLJ-GEN-2026-000001')
    include_qr = data.get('include_qr', True)
    org = OrgSettings.query.first() or OrgSettings()

    mode = ct.render_mode or ('overlay' if ct.master_pdf_path else 'mlj')
    if mode == 'overlay' and not os.path.exists(ct.master_pdf_path or ''):
        return jsonify({'error': 'Master certificate file not found on server. Please re-upload the template.'}), 404

    preview_user = type('PreviewUser', (), {
        'full_name': full_name,
        'certificate_id': cert_id,
        'include_qr': include_qr,
        'sent_at': None,
        'approved_at': None,
    })()

    try:
        from app.engine.renderer import render_certificate_pdf
        pdf_bytes = render_certificate_pdf(preview_user, ct, org)
    except FileNotFoundError:
        return jsonify({'error': 'Master certificate file missing.'}), 404
    except Exception as e:
        return jsonify({'error': f'Failed to generate preview: {str(e)}'}), 500

    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=False,
        download_name=f"preview_{ct_id}.pdf"
    )

@bp.route('/api/users/<int:cert_type_id>')
@login_required
def get_users(cert_type_id):
    status_filter = request.args.get('status')
    q = User.query.filter_by(certificate_type_id=cert_type_id)
    if status_filter:
        q = q.filter_by(status=status_filter)
    users = q.order_by(User.created_at.desc()).all()
    return jsonify({'data': [{
        'id': u.id,
        'full_name': u.full_name,
        'email': u.email,
        'certificate_id': u.certificate_id or '—',
        'status': u.status,
        'include_qr': u.include_qr,
        'score': u.score,
        'source': u.source,
        'generated': u.generated_at is not None,
        'created_at': u.created_at.strftime('%d/%m/%Y %H:%M') if u.created_at else '',
    } for u in users]})


@bp.route('/api/users/add', methods=['POST'])
@login_required
def add_user():
    data = request.json or {}
    ct_id = data.get('cert_type_id')
    if not ct_id:
        return jsonify({'error': 'cert_type_id required'}), 400
    ct = CertificateType.query.get_or_404(ct_id)

    first_name = (data.get('first_name') or '').strip()
    surname    = (data.get('surname')    or '').strip()
    email      = (data.get('email')      or '').strip().lower()
    if not first_name or not surname or not email:
        return jsonify({'error': 'first_name, surname and email required'}), 400

    seq     = next_sequence(ct)
    cert_id = generate_cert_id(ct.course_code, seq, datetime.utcnow())

    user = User(
        first_name=first_name,
        surname=surname,
        other_name=(data.get('other_name') or '').strip(),
        email=email,
        certificate_type_id=ct_id,
        certificate_id=cert_id,
        status='registered',
        source='manual',
    )
    db.session.add(user)
    db.session.add(AuditLog(action='registered', performed_by=current_user.email,
                            details={'cert_id': cert_id, 'email': email}))
    db.session.commit()
    return jsonify({'message': 'Participant added', 'id': user.id, 'certificate_id': cert_id})


@bp.route('/api/users/import', methods=['POST'])
@login_required
def import_users():
    import openpyxl
    ct_id = request.form.get('cert_type_id')
    ct = CertificateType.query.get_or_404(ct_id)
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'File required'}), 400

    ext = file.filename.rsplit('.', 1)[-1].lower()
    rows = []

    if ext == 'csv':
        content = file.read().decode('utf-8', errors='ignore')
        reader = csv.DictReader(io.StringIO(content))
        rows = list(reader)
    elif ext in ('xlsx', 'xls'):
        wb = openpyxl.load_workbook(io.BytesIO(file.read()))
        ws = wb.active
        headers = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
    else:
        return jsonify({'error': 'CSV or XLSX only'}), 400

    imported = skipped = 0
    for row in rows:
        email = str(row.get('email') or row.get('Email') or '').strip().lower()
        first_name = str(row.get('first_name') or row.get('First Name') or row.get('firstname') or '').strip()
        surname = str(row.get('surname') or row.get('last_name') or row.get('Last Name') or row.get('lastname') or '').strip()
        if not email or not first_name:
            skipped += 1
            continue

        if not first_name and not surname:
            full = str(row.get('full_name') or row.get('name') or '')
            parts = full.split()
            first_name = parts[0] if parts else ''
            surname = parts[-1] if len(parts) > 1 else ''

        seq = next_sequence(ct)
        cert_id = generate_cert_id(ct.course_code, seq, datetime.utcnow())

        user = User(
            first_name=first_name, surname=surname,
            email=email,
            certificate_type_id=int(ct_id),
            certificate_id=cert_id,
            status='registered',
            score=float(row.get('score') or 0) if row.get('score') else None,
            completion_status=str(row.get('completion_status') or '').strip() or None,
            source='csv',
        )
        db.session.add(user)
        imported += 1

    db.session.commit()
    return jsonify({'message': f'{imported} imported, {skipped} skipped'})


@bp.route('/api/users/toggle-qr', methods=['POST'])
@login_required
def toggle_qr():
    uid = request.json.get('user_id')
    user = db.session.get(User, uid)
    if not user:
        return jsonify({'error': 'Not found'}), 404
    user.include_qr = not user.include_qr
    db.session.commit()
    return jsonify({'include_qr': user.include_qr})


@bp.route('/api/approve', methods=['POST'])
@login_required
def approve_users():
    """Approve (single or batch) after manual payment verification.
    Approval immediately generates each registrant's certificate in the
    background and archives the PDF — dispatch remains a separate,
    admin-triggered step."""
    from app.services.email.queue import run_in_background
    from app.services.certgen import generate_certificates_job

    user_ids = request.json['user_ids']
    eligible = [u.id for u in User.query.filter(
        User.id.in_(user_ids),
        User.status.in_(['registered', 'rejected'])).all()]

    User.query.filter(User.id.in_(eligible)).update(
        {'status': 'approved', 'approved_at': datetime.utcnow()},
        synchronize_session=False)
    for uid in eligible:
        db.session.add(AuditLog(user_id=uid, action='approved',
                                performed_by=current_user.email))
    db.session.commit()

    if eligible:
        run_in_background(generate_certificates_job, user_ids=eligible)

    return jsonify({
        'message': f'{len(eligible)} approved — certificates are being generated',
        'approved': len(eligible),
    })


@bp.route('/api/reject', methods=['POST'])
@login_required
def reject_users():
    user_ids = request.json['user_ids']
    User.query.filter(User.id.in_(user_ids)).update(
        {'status': 'rejected'}, synchronize_session=False)
    db.session.commit()
    return jsonify({'message': f'{len(user_ids)} rejected'})


@bp.route('/api/send', methods=['POST'])
@login_required
def send_certificates():
    data = request.json or {}
    user_ids = data['user_ids']
    cert_type_id = data['cert_type_id']
    draft_id = data.get('draft_id')
    confirmed = data.get('confirmed', False)

    if not confirmed:
        users = User.query.filter(
            User.id.in_(user_ids), User.status == 'approved').all()
        ct = CertificateType.query.get(cert_type_id)
        org = OrgSettings.query.first() or OrgSettings()
        return jsonify({
            'needs_confirmation': True,
            'recipient_count': len(users),
            'course': ct.name if ct else '',
            'sender': f"{org.sender_name} <{org.sender_email or current_app.config.get('MAIL_USERNAME','')}>",
            'cert_type_id': cert_type_id,
        })

    from app.services.email.queue import enqueue_batch

    users = User.query.filter(
        User.id.in_(user_ids), User.status == 'approved').all()
    queued_ids = []
    for user in users:
        user.status = 'sending'
        user.sent_at = datetime.utcnow()
        queued_ids.append(user.id)

    db.session.add(AuditLog(
        action='batch_send_initiated', performed_by=current_user.email,
        details={'count': len(queued_ids), 'cert_type_id': cert_type_id}
    ))
    db.session.commit()

    if queued_ids:
        enqueue_batch(queued_ids, cert_type_id, draft_id)

    return jsonify({'message': f'{len(queued_ids)} certificates queued for dispatch'})


@bp.route('/api/send-all-approved', methods=['POST'])
@login_required
def send_all_approved():
    data = request.json or {}
    cert_type_id = data['cert_type_id']
    draft_id = data.get('draft_id')
    confirmed = data.get('confirmed', False)

    users = User.query.filter_by(
        certificate_type_id=cert_type_id, status='approved').all()

    if not confirmed:
        ct = CertificateType.query.get(cert_type_id)
        org = OrgSettings.query.first() or OrgSettings()
        return jsonify({
            'needs_confirmation': True,
            'recipient_count': len(users),
            'course': ct.name if ct else '',
            'sender': f"{org.sender_name} <{org.sender_email or current_app.config.get('MAIL_USERNAME','')}>",
        })

    if not users:
        return jsonify({'message': 'No approved participants found'})

    from app.services.email.queue import enqueue_batch

    queued_ids = []
    for user in users:
        user.status = 'sending'
        user.sent_at = datetime.utcnow()
        queued_ids.append(user.id)

    db.session.add(AuditLog(
        action='send_all_approved', performed_by=current_user.email,
        details={'count': len(queued_ids), 'cert_type_id': cert_type_id}
    ))
    db.session.commit()

    enqueue_batch(queued_ids, cert_type_id, draft_id)

    return jsonify({'message': f'{len(queued_ids)} certificates queued for dispatch'})


@bp.route('/api/nudge', methods=['POST'])
@login_required
def nudge_user():
    from app.services.email.queue import run_in_background
    from app.worker import send_nudge_email

    uid = request.json['user_id']
    run_in_background(send_nudge_email, user_id=uid)
    return jsonify({'message': 'Reminder queued'})


@bp.route('/api/users/<int:uid>/certificate.pdf')
@login_required
def download_user_certificate(uid):
    """View/download a registrant's generated certificate. Generates on
    demand if the registrant is approved but the archive copy is missing."""
    from flask import send_file
    from app.services.certgen import generate_for_user

    user = db.session.get(User, uid)
    if not user:
        return jsonify({'error': 'Not found'}), 404

    archive = CertArchive.query.filter_by(certificate_id=user.certificate_id).first()
    pdf_bytes = archive.pdf_binary if archive and archive.pdf_binary else None

    if not pdf_bytes:
        if user.status not in ('approved', 'sending', 'sent'):
            return jsonify({'error': 'Certificate not generated — registrant is not approved yet'}), 409
        cert_type = db.session.get(CertificateType, user.certificate_type_id)
        org = OrgSettings.query.first() or OrgSettings()
        try:
            pdf_bytes = generate_for_user(user, cert_type, org)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f'Generation failed: {e}'}), 500

    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=request.args.get('download') == '1',
        download_name=f"Certificate_{user.certificate_id}.pdf",
    )


@bp.route('/api/users/<int:uid>/regenerate', methods=['POST'])
@login_required
def regenerate_user_certificate(uid):
    """Re-render an approved registrant's certificate (e.g. after fixing a
    name) and replace the archived copy."""
    from app.services.certgen import generate_for_user

    user = db.session.get(User, uid)
    if not user:
        return jsonify({'error': 'Not found'}), 404
    if user.status not in ('approved', 'sending', 'sent'):
        return jsonify({'error': 'Only approved registrants have certificates'}), 409

    cert_type = db.session.get(CertificateType, user.certificate_type_id)
    org = OrgSettings.query.first() or OrgSettings()
    try:
        generate_for_user(user, cert_type, org)
        db.session.add(AuditLog(user_id=uid, action='regenerated',
                                performed_by=current_user.email,
                                details={'cert_id': user.certificate_id}))
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Generation failed: {e}'}), 500
    return jsonify({'message': 'Certificate regenerated'})


@bp.route('/api/revoke', methods=['POST'])
@login_required
def revoke_certificates():
    """Revoke issued certificates — mirrors the MedLocum revocation flow.
    Revoked certificates fail public verification."""
    user_ids = request.json['user_ids']
    users = User.query.filter(User.id.in_(user_ids)).all()
    revoked = 0
    for u in users:
        if not u.certificate_id:
            continue
        u.revoked_at = datetime.utcnow()
        u.status = 'revoked'
        archive = CertArchive.query.filter_by(certificate_id=u.certificate_id).first()
        if archive:
            archive.status = 'revoked'
        db.session.add(AuditLog(user_id=u.id, action='revoked',
                                performed_by=current_user.email,
                                details={'cert_id': u.certificate_id}))
        revoked += 1
    db.session.commit()
    return jsonify({'message': f'{revoked} certificates revoked'})


@bp.route('/api/delete', methods=['POST'])
@login_required
def delete_users():
    user_ids = request.json['user_ids']
    users = User.query.filter(User.id.in_(user_ids)).all()
    for u in users:
        if u.certificate_id and not CertArchive.query.filter_by(
                certificate_id=u.certificate_id).first():
            db.session.add(CertArchive(
                certificate_id=u.certificate_id,
                full_name=u.full_name,
                cert_name=u.certificate_type.name if u.certificate_type else '',
                course_code=u.certificate_type.course_code if u.certificate_type else '',
                issued_date=u.sent_at.strftime('%d %B %Y') if u.sent_at else '',
                email=u.email,
                status='archived',
                raw_binary=json.dumps({'email': u.email}).encode('utf-8')
            ))
    User.query.filter(User.id.in_(user_ids)).delete(synchronize_session=False)
    db.session.add(AuditLog(
        action='deleted', performed_by=current_user.email,
        details={'count': len(user_ids)}
    ))
    db.session.commit()
    return jsonify({'message': f'{len(user_ids)} participants deleted'})


@bp.route('/api/users/archive', methods=['POST'])
@login_required
def archive_users():
    user_ids = request.json['user_ids']
    User.query.filter(User.id.in_(user_ids)).update(
        {'archived_at': datetime.utcnow(), 'status': 'archived'},
        synchronize_session=False)
    db.session.commit()
    return jsonify({'message': f'{len(user_ids)} participants archived'})

@bp.route('/api/drafts', methods=['GET'])
@login_required
def list_drafts():
    drafts = EmailDraft.query.order_by(EmailDraft.updated_at.desc()).all()
    return jsonify([{
        'id': d.id, 'name': d.name, 'subject': d.subject,
        'body': d.body, 'include_attachment': d.include_attachment,
        'is_default': d.is_default,
        'updated_at': d.updated_at.strftime('%d/%m/%Y') if d.updated_at else ''
    } for d in drafts])


@bp.route('/api/drafts', methods=['POST'])
@login_required
def save_draft():
    data = request.json or {}
    draft = EmailDraft(
        name=data.get('name', 'Untitled Draft'),
        subject=data.get('subject', 'Your {{course_name}} Certificate'),
        body=data.get('body', ''),
        include_attachment=data.get('include_attachment', True),
        is_default=data.get('is_default', False),
    )
    db.session.add(draft)
    db.session.commit()
    return jsonify({'id': draft.id, 'message': 'Draft saved'})


@bp.route('/api/drafts/<int:did>', methods=['PUT'])
@login_required
def update_draft(did):
    draft = db.session.get(EmailDraft, did)
    if not draft:
        return jsonify({'error': 'Not found'}), 404
    data = request.json or {}
    for f in ('name', 'subject', 'body', 'include_attachment', 'is_default'):
        if f in data:
            setattr(draft, f, data[f])
    draft.updated_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'message': 'Draft updated'})


@bp.route('/api/drafts/<int:did>', methods=['DELETE'])
@login_required
def delete_draft(did):
    draft = db.session.get(EmailDraft, did)
    if not draft:
        return jsonify({'error': 'Not found'}), 404
    db.session.delete(draft)
    db.session.commit()
    return jsonify({'message': 'Draft deleted'})

@bp.route('/api/campaigns', methods=['GET'])
@login_required
def list_campaigns():
    camps = Campaign.query.order_by(Campaign.created_at.desc()).all()
    return jsonify([{
        'id': c.id, 'name': c.name, 'status': c.status,
        'recipient_count': c.recipient_count,
        'sent_count': c.sent_count, 'failed_count': c.failed_count,
        'include_attachment': c.include_attachment,
        'created_at': c.created_at.strftime('%d/%m/%Y') if c.created_at else '',
        'sent_at': c.sent_at.strftime('%d/%m/%Y %H:%M') if c.sent_at else '',
    } for c in camps])


@bp.route('/api/campaigns', methods=['POST'])
@login_required
def create_campaign():
    data = request.json or {}
    user_ids = data.get('user_ids', [])
    draft_id = data.get('draft_id')
    cert_type_id = data.get('cert_type_id')
    send_now = data.get('send_now', False)

    campaign = Campaign(
        name=data.get('name', f"Campaign {datetime.utcnow().strftime('%d/%m/%Y')}"),
        cert_type_id=cert_type_id,
        draft_id=draft_id,
        include_attachment=data.get('include_attachment', False),
        recipient_count=len(user_ids),
        status='draft',
        created_by=current_user.email,
    )
    db.session.add(campaign)
    db.session.commit()

    if send_now:
        from app.services.email.queue import run_in_background
        from app.worker import process_campaign

        run_in_background(
            process_campaign,
            campaign_id=campaign.id,
            user_ids=user_ids,
            draft_id=draft_id,
        )
        campaign.status = 'scheduled'
        db.session.commit()

    return jsonify({'id': campaign.id, 'message': 'Campaign created'})

@bp.route('/api/lms/completion', methods=['POST'])
def lms_completion():
    data = request.json or {}
    email = (data.get('email') or data.get('learner_email') or '').strip().lower()
    full_name = (data.get('full_name') or data.get('name') or '').strip()
    score = data.get('score') or data.get('grade')
    completion = data.get('completion_status') or data.get('status') or 'completed'
    token = data.get('registration_token') or request.args.get('token', '')

    if not email or not token:
        return jsonify({'error': 'email and registration_token required'}), 400

    ct = CertificateType.query.filter_by(registration_token=token, is_active=True).first()
    if not ct:
        return jsonify({'error': 'Invalid token'}), 404

    name_parts = full_name.split()
    if len(name_parts) == 1:
        first_name = name_parts[0]
        surname = ""
    elif len(name_parts) > 1:
        first_name = name_parts[0]
        surname = " ".join(name_parts[1:])
    else:
        first_name = email.split('@')[0]
        surname = ""

    seq = next_sequence(ct)
    cert_id = generate_cert_id(ct.course_code, seq, datetime.utcnow())

    user = User(
        first_name=first_name,
        surname=surname,
        email=email,
        certificate_type_id=ct.id,
        certificate_id=cert_id,
        status='registered',
        score=float(score) if score else None,
        completion_status=completion,
        source='lms',
    )
    db.session.add(user)
    db.session.commit()

    return jsonify({
        'message': 'Participant registered',
        'certificate_id': cert_id,
        'status': 'pending_admin_approval'
    })


@bp.route('/api/lms/users', methods=['POST'])
def lms_users():
    data = request.json or {}
    token = data.get('registration_token') or request.args.get('token', '')
    users_data = data.get('users', [])

    ct = CertificateType.query.filter_by(registration_token=token, is_active=True).first()
    if not ct:
        return jsonify({'error': 'Invalid token'}), 404

    imported = 0
    for u in users_data:
        email = (u.get('email') or '').strip().lower()
        if not email:
            continue
        full = (u.get('full_name') or u.get('name') or '').strip()
        name_parts = full.split()
        if len(name_parts) == 1:
            f_name = name_parts[0]
            s_name = ""
        elif len(name_parts) > 1:
            f_name = name_parts[0]
            s_name = " ".join(name_parts[1:])
        else:
            f_name = email.split('@')[0]
            s_name = ""

        seq = next_sequence(ct)
        cert_id = generate_cert_id(ct.course_code, seq, datetime.utcnow())

        user = User(
            first_name=f_name,
            surname=s_name,
            email=email,
            certificate_type_id=ct.id,
            certificate_id=cert_id,
            status='registered',
            source='lms',
        )
        db.session.add(user)
        imported += 1

    db.session.commit()
    return jsonify({'message': f'{imported} users synced'})

@bp.route('/verify/<cert_id>')
def verify_cert(cert_id):
    """Public certificate verification (the QR code target). Mirrors the
    MedLocum semantics: a certificate verifies only if it exists and has
    not been revoked."""
    revoked = False
    record = CertArchive.query.filter_by(certificate_id=cert_id).first()
    if record and record.status == 'revoked':
        revoked = True
    if not record:
        user = User.query.filter_by(certificate_id=cert_id).first()
        if user:
            revoked = user.revoked_at is not None or user.status == 'revoked'
            record = type('R', (), {
                'certificate_id': user.certificate_id,
                'full_name': user.full_name,
                'cert_name': user.certificate_type.name if user.certificate_type else '',
                'issued_date': (user.sent_at or user.approved_at).strftime('%d %B %Y')
                               if (user.sent_at or user.approved_at) else '',
                'status': user.status,
            })()
    if revoked:
        record = None
    fmt_valid = verify_format(cert_id)
    return render_template('public/verify.html',
                           record=record, fmt_valid=fmt_valid,
                           revoked=revoked, cert_id=cert_id)


# unsubscribe route lives in email_routes.py to avoid duplicate endpoint
@bp.route('/api/archive/<cert_id>')
@login_required
def get_archive(cert_id):
    record = CertArchive.query.filter_by(certificate_id=cert_id).first()
    if not record:
        return jsonify({'error': 'Not found'}), 404
    data = record.to_dict()
    if record.raw_binary:
        try:
            data['raw'] = json.loads(record.raw_binary.decode('utf-8'))
        except Exception:
            pass
    return jsonify(data)

@bp.route('/api/audit')
@login_required
def get_audit():
    logs = AuditLog.query.order_by(AuditLog.created_at.desc()).limit(200).all()
    return jsonify([{
        'action': l.action,
        'performed_by': l.performed_by,
        'details': l.details,
        'created_at': l.created_at.strftime('%d/%m/%Y %H:%M:%S') if l.created_at else '',
    } for l in logs])
